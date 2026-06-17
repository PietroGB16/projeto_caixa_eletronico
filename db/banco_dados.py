"""
banco_dados.py

Camada de PERSISTENCIA. Usa uma planilha Excel (.xlsx) como banco de
dados das movimentacoes.

Esta camada nao faz contas (isso e do dominio, cliente_bancario.py) nem
fala com o usuario (isso e da interface, main.py). Ela so sabe duas
coisas: ler a planilha e gravar a planilha.

Estrutura da planilha (uma linha por movimentacao):
    Tipo | Valor | Data
"""

import os
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


class RepositorioExcel:

    CABECALHO = ["Tipo", "Valor", "Data"]

    def __init__(self, caminho="banco_dados.xlsx"):
        self.caminho = caminho

    def existe(self):
        """Diz se o arquivo da planilha ja existe."""
        return os.path.exists(self.caminho)

    # ------------------------------------------------------------------
    def carregar(self):
        """Le a planilha e devolve a lista de movimentacoes no mesmo
        formato que o dominio usa: dicts com Tipo/Valor/Data."""
        wb = load_workbook(self.caminho)
        sheet = wb.active

        movimentacoes = []
        # min_row=2 pula a linha do cabecalho
        for linha in sheet.iter_rows(min_row=2, values_only=True):
            tipo, valor, data = linha
            if tipo is None:          # ignora linhas em branco
                continue
            movimentacoes.append({
                "Tipo": tipo,
                "Valor": float(valor),
                "Data": self._para_date(data),
            })
        return movimentacoes

    # ------------------------------------------------------------------
    def salvar(self, movimentacoes):
        """Grava a lista inteira de movimentacoes (sobrescreve o arquivo).

        Estrategia simples: a cada operacao a interface manda salvar tudo
        de novo. Para um trabalho academico e mais que suficiente.
        """
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Movimentacoes"

        # Cabecalho em negrito
        sheet.append(self.CABECALHO)
        for celula in sheet[1]:
            celula.font = Font(bold=True, name="Arial")

        # Uma linha por movimentacao
        for mov in movimentacoes:
            sheet.append([mov["Tipo"], mov["Valor"], mov["Data"]])

        # Coluna de data no formato brasileiro
        for linha in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            for celula in linha:
                celula.number_format = "DD/MM/YYYY"

        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 14

        wb.save(self.caminho)

    # ------------------------------------------------------------------
    @staticmethod
    def _para_date(valor):
        """Converte o que a planilha devolveu em um date do Python.

        O Excel costuma devolver datas como datetime; aqui normalizamos
        para date (e tratamos tambem o caso de ter vindo como texto).
        """
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        return datetime.strptime(str(valor), "%d/%m/%Y").date()